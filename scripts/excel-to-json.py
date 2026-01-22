#!/usr/bin/env python3
"""
Script para convertir microsite-faqs.xlsx a JSON para la aplicación de demo.

Uso:
    python3 scripts/excel-to-json.py

Este script lee el Excel con las conversaciones de cada tipo de viajero
y genera un archivo JSON que puede ser importado en la aplicación.
"""

import openpyxl
import json
import os
from pathlib import Path

# Configuración
EXCEL_PATH = 'doc/microsite-faqs.xlsx'
OUTPUT_JSON = 'src/app/demo/conversations-data.json'
OUTPUT_TS = 'src/app/demo/conversations-data.ts'

# Mapeo de perfiles a iconos y descripciones
PROFILE_CONFIG = {
    'Trabajo Solo': {
        'icon': '💼',
        'description': 'Viajero de negocios individual',
        'caracteristica': 'trabajo',
        'grupo': 'solo'
    },
    'Trabajo Pareja': {
        'icon': '💼👥',
        'description': 'Viajeros de negocios en pareja',
        'caracteristica': 'trabajo',
        'grupo': 'pareja'
    },
    'Trabajo Grupo': {
        'icon': '💼👨‍👩‍👧‍👦',
        'description': 'Grupo de negocios o equipo de trabajo',
        'caracteristica': 'trabajo',
        'grupo': 'grupo'
    },
    'Descanso Solo': {
        'icon': '🌴',
        'description': 'Viajero individual buscando relajación',
        'caracteristica': 'descanso',
        'grupo': 'solo'
    },
    'Descanso Pareja': {
        'icon': '🌴👥',
        'description': 'Pareja buscando relax y desconexión',
        'caracteristica': 'descanso',
        'grupo': 'pareja'
    },
    'Descanso Grupo': {
        'icon': '🌴👨‍👩‍👧‍👦',
        'description': 'Grupo de amigos o familia en modo relax',
        'caracteristica': 'descanso',
        'grupo': 'grupo'
    },
    'Aventura Solo': {
        'icon': '🧭',
        'description': 'Viajero aventurero explorando solo',
        'caracteristica': 'aventura',
        'grupo': 'solo'
    },
    'Aventura Pareja': {
        'icon': '🧭👥',
        'description': 'Pareja de aventureros explorando juntos',
        'caracteristica': 'aventura',
        'grupo': 'pareja'
    },
    'Aventura Grupo': {
        'icon': '🧭👨‍👩‍👧‍👦',
        'description': 'Grupo de aventureros explorando Lima',
        'caracteristica': 'aventura',
        'grupo': 'grupo'
    }
}


def extract_content_blocks(sheet):
    """
    Extrae bloques de contenido de una pestaña del Excel.

    Cada bloque tiene la siguiente estructura vertical:
    - Fila N: TOPIC
    - Fila N+1: SUB TEMAS
    - Fila N+2: Texto intro gente
    - Fila N+3: Fase Categoría
    - Fila N+4: Título Unidad de Contenido
    - Fila N+5: Párrafo / Contenido descripción
    - Fila N+6: Imagen / Video / Multimedia
    - Fila N+7: Texto cierre agente
    - Fila N+8: Texto agente próximo paso
    - Fila N+9: CTAS Recomendados
    """
    blocks = []

    # Buscar TODOS los bloques en todas las columnas y filas
    for col in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):  # Buscar en todas las filas
            cell_value = sheet.cell(row, col).value

            # Identificar inicio de bloque (celda con "TOPIC")
            if cell_value == "TOPIC":
                topic = sheet.cell(row, col + 1).value
                sub_tema = sheet.cell(row + 1, col + 1).value

                # Solo extraer si tiene al menos topic y subtema
                if topic or sub_tema:
                    block = {
                        'topic': topic or '',
                        'sub_tema': sub_tema or '',
                        'intro': sheet.cell(row + 2, col + 1).value or '',
                        'fase': sheet.cell(row + 3, col + 1).value or '',
                        'titulo': sheet.cell(row + 4, col + 1).value or '',
                        'contenido': sheet.cell(row + 5, col + 1).value or '',
                        'imagen': sheet.cell(row + 6, col + 1).value or '',
                        'cierre': sheet.cell(row + 7, col + 1).value or '',
                        'proximo_paso': sheet.cell(row + 8, col + 1).value or '',
                        'ctas': sheet.cell(row + 9, col + 1).value or '',
                    }

                    # Solo agregar si tiene contenido significativo
                    if block['titulo'] and block['contenido']:
                        blocks.append(block)

    return blocks


def convert_excel_to_json():
    """Convierte el Excel completo a estructura JSON."""

    # Verificar que existe el archivo Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: No se encontró el archivo {EXCEL_PATH}")
        return None

    print(f"📖 Leyendo {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    profiles = []

    # Procesar cada pestaña de perfil
    for sheet_name, config in PROFILE_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Advertencia: Pestaña '{sheet_name}' no encontrada")
            continue

        sheet = wb[sheet_name]
        blocks = extract_content_blocks(sheet)

        profile = {
            'name': sheet_name,
            'icon': config['icon'],
            'description': config['description'],
            'caracteristica': config['caracteristica'],
            'grupo': config['grupo'],
            'conversations': blocks
        }

        profiles.append(profile)
        print(f"  ✓ {sheet_name}: {len(blocks)} conversaciones extraídas")

    print(f"\n✅ Total: {len(profiles)} perfiles procesados")
    return profiles


def save_json(data, filepath):
    """Guarda los datos en formato JSON."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"💾 JSON guardado en: {filepath}")


def generate_typescript_file(data, filepath):
    """Genera el archivo TypeScript con los tipos e interfaces."""

    ts_content = '''// Archivo generado automáticamente desde microsite-faqs.xlsx
// Para regenerar: python3 scripts/excel-to-json.py

export interface Conversation {
  topic: string
  sub_tema: string
  intro?: string
  fase: string
  titulo: string
  contenido: string
  imagen?: string
  cierre?: string
  proximo_paso?: string
  ctas: string
}

export interface Profile {
  name: string
  icon: string
  description: string
  caracteristica: string
  grupo: string
  conversations: Conversation[]
}

export const profiles: Profile[] = ''' + json.dumps(data, ensure_ascii=False, indent=2) + '\n'

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(ts_content)

    print(f"💾 TypeScript guardado en: {filepath}")


def main():
    """Función principal."""
    print("=" * 60)
    print("  CONVERSIÓN EXCEL → JSON/TypeScript")
    print("=" * 60)
    print()

    # Convertir Excel a datos
    data = convert_excel_to_json()

    if not data:
        print("❌ Error en la conversión")
        return

    print()

    # Guardar JSON
    save_json(data, OUTPUT_JSON)

    # Generar TypeScript
    generate_typescript_file(data, OUTPUT_TS)

    print()
    print("=" * 60)
    print("✅ CONVERSIÓN COMPLETADA")
    print("=" * 60)
    print()
    print("📝 Próximos pasos:")
    print("  1. Revisar el archivo generado en src/app/demo/conversations-data.ts")
    print("  2. La aplicación /demo ahora tiene los 9 perfiles completos")
    print("  3. Puedes editar el Excel y volver a ejecutar este script")
    print()


if __name__ == '__main__':
    main()
